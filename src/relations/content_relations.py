from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import json

import numpy as np
import yaml

from src.config import settings
from src.indexing.embedder import Embedder
from src.indexing.extractor import extract_text, list_archive_members
from src.models.descriptor import FileDescriptor
from src.relations.base import RelationEdge, RelationParser
from src.relations.version_relations import _strip_md_wrapper
from typing import Any as Neo4jStore

IMPORT_PATTERNS = [
    (re.compile(r"^\s*import\s+([\w.]+)", re.M), ".py"),
    (re.compile(r"^\s*from\s+([\w.]+)\s+import", re.M), ".py"),
    (re.compile(r'#include\s*[<"]([^>"]+)[>"]', re.M), None),
    (re.compile(r"""require\s*\(\s*['"]([^'"]+)['"]\s*\)""", re.M), ".js"),
]
REF_PATTERNS = [
    re.compile(r"\[([^\]]*)\]\(file://([^)]+)\)", re.I),
    re.compile(r"file://([^\s\)\]\"']+)", re.I),
    # 邮件/文档正文中的文件名引用（HippoCamp 等个人文件集）
    re.compile(
        r"([A-Za-z0-9][A-Za-z0-9_\-\(\)% ]{0,120}\.(?:pdf|docx?|xlsx?|pptx?|eml|ics|md|txt|csv|png|jpe?g))",
        re.I,
    ),
]
YAML_REF_PATTERNS = [
    re.compile(r"\$ref\s*:\s*['\"]?([^'\"\n]+)", re.I),
    re.compile(r"!include\s+['\"]?([^'\"\n]+)", re.I),
]
JSON_REF_PATTERNS = [
    re.compile(r'"?\$ref"?\s*:\s*"([^"]+)"', re.I),
    re.compile(r'"include"\s*:\s*"([^"]+)"', re.I),
]
TEMP_PATTERNS = [
    (re.compile(r"^~\$(.+)$", re.I), "office_temp"),
    (re.compile(r"^\.~(.+)$"), "swp"),
    (re.compile(r"(.+)\.tmp$", re.I), "tmp"),
]
BACKUP_PATTERNS = [
    re.compile(r"^(.+)_backup(\.\w+)$", re.I),
    re.compile(r"^(.+)\.bak$", re.I),
    re.compile(r"^(.+)\.backup(\.\w+)?$", re.I),
]


class SimilarToParser(RelationParser):
    name = "similar_to"

    def discover(
        self, descriptors: list[FileDescriptor], store: Neo4jStore
    ) -> list[RelationEdge]:
        valid = [
            d
            for d in descriptors
            if d.file_embedding
            and d.status.value == "ACTIVE"
            and "/noise/" not in d.path.replace("\\", "/").lower()
        ]
        if len(valid) < 2:
            return []

        matrix = np.array([d.file_embedding for d in valid], dtype=np.float32)
        norms = np.linalg.norm(matrix, axis=1, keepdims=True) + 1e-9
        matrix = matrix / norms

        edges: list[RelationEdge] = []
        threshold = settings.similar_threshold
        top_k = settings.similar_top_k

        try:
            import faiss

            index = faiss.IndexFlatIP(matrix.shape[1])
            index.add(matrix)
            for i, d in enumerate(valid):
                D, I = index.search(matrix[i : i + 1], min(top_k + 1, len(valid)))
                for j, score in zip(I[0], D[0]):
                    if j < 0 or j == i or score < threshold:
                        continue
                    if i < j:
                        edges.append(
                            RelationEdge(
                                d.file_id,
                                "SIMILAR_TO",
                                valid[j].file_id,
                                weight=float(score),
                                symmetric=True,
                                props={"similarity": float(score), "index": "faiss"},
                            )
                        )
            return edges
        except ImportError:
            sims = matrix @ matrix.T
            np.fill_diagonal(sims, 0)
            for i, d in enumerate(valid):
                scores = sims[i]
                candidates = np.where(scores >= threshold)[0]
                ranked = sorted(candidates, key=lambda j: scores[j], reverse=True)[:top_k]
                for j in ranked:
                    if i < j:
                        edges.append(
                            RelationEdge(
                                d.file_id,
                                "SIMILAR_TO",
                                valid[j].file_id,
                                weight=float(scores[j]),
                                symmetric=True,
                                props={"similarity": float(scores[j])},
                            )
                        )
            return edges


class DependsOnParser(RelationParser):
    name = "depends_on"

    def discover(
        self, descriptors: list[FileDescriptor], store: Neo4jStore
    ) -> list[RelationEdge]:
        path_map: dict[str, FileDescriptor] = {}
        name_map: dict[str, list[FileDescriptor]] = defaultdict(list)
        for d in descriptors:
            path_map[Path(d.path).resolve().as_posix().lower()] = d
            name_map[d.name.lower()].append(d)

        edges: list[RelationEdge] = []
        code_exts = {".py", ".js", ".ts", ".c", ".cpp", ".h", ".java"}
        config_exts = {".yaml", ".yml", ".json"}

        for d in descriptors:
            if "/noise/" in d.path.replace("\\", "/").lower():
                continue
            try:
                text = extract_text(Path(d.path), 50000)
            except Exception:
                continue
            ext = d.extension
            if ext == ".md" and "." in d.name:
                inner = Path(_strip_md_wrapper(d.name)).suffix.lower()
                if inner in code_exts:
                    ext = inner
            is_code = ext in code_exts or (
                d.extension == ".md" and re.search(r"^\s*(import|from)\s+", text, re.M)
            )
            if not is_code and ext not in config_exts:
                continue
            base_dir = Path(d.path).parent
            refs = self._parse_imports(text, ".py") if is_code else []
            if ext in config_exts:
                refs.extend(self._parse_config_refs(Path(d.path), text))
            for ref in refs:
                target = self._resolve_ref(ref, base_dir, path_map, name_map)
                if target and target.file_id != d.file_id:
                    edges.append(
                        RelationEdge(d.file_id, "DEPENDS_ON", target.file_id, weight=0.9)
                    )
        return edges

    def _parse_imports(self, text: str, ext: str) -> list[str]:
        refs: list[str] = []
        for pattern, required_ext in IMPORT_PATTERNS:
            if required_ext and ext != required_ext:
                continue
            for m in pattern.finditer(text):
                refs.append(m.group(1))
        return refs

    def _parse_config_refs(self, path: Path, text: str) -> list[str]:
        refs: list[str] = []
        for pat in YAML_REF_PATTERNS + JSON_REF_PATTERNS:
            refs.extend(pat.findall(text))
        if path.suffix.lower() in (".yaml", ".yml"):
            try:
                data = yaml.safe_load(text)
                refs.extend(self._walk_refs(data))
            except Exception:
                pass
        elif path.suffix.lower() == ".json":
            try:
                data = json.loads(text)
                refs.extend(self._walk_refs(data))
            except Exception:
                pass
        return refs

    def _walk_refs(self, obj, depth: int = 0) -> list[str]:
        if depth > 8:
            return []
        out: list[str] = []
        if isinstance(obj, dict):
            if "$ref" in obj:
                out.append(str(obj["$ref"]))
            if "include" in obj:
                out.append(str(obj["include"]))
            for v in obj.values():
                out.extend(self._walk_refs(v, depth + 1))
        elif isinstance(obj, list):
            for v in obj:
                out.extend(self._walk_refs(v, depth + 1))
        return out

    def _resolve_ref(
        self,
        ref: str,
        base_dir: Path,
        path_map: dict,
        name_map: dict,
    ) -> FileDescriptor | None:
        ref = ref.replace(".", "/").strip("/")
        candidates = [
            base_dir / ref,
            base_dir / f"{ref}.py",
            base_dir / f"{ref}.js",
            base_dir / f"{ref}.ts",
        ]
        for c in candidates:
            key = c.resolve().as_posix().lower()
            if key in path_map:
                return path_map[key]
        name = Path(ref).name.lower()
        if name in name_map:
            return name_map[name][0]
        stem = name.replace(".py", "").replace(".js", "")
        for key, descs in name_map.items():
            kl = key.lower()
            if kl.startswith(stem) and any(x in kl for x in (".py", ".js", ".yaml", ".md")):
                return descs[0]
        return None


class ReferencesParser(RelationParser):
    name = "references"

    def discover(
        self, descriptors: list[FileDescriptor], store: Neo4jStore
    ) -> list[RelationEdge]:
        path_map = {Path(d.path).resolve().as_posix().lower(): d for d in descriptors}
        name_map: dict[str, FileDescriptor] = {d.name.lower(): d for d in descriptors}
        edges: list[RelationEdge] = []

        for d in descriptors:
            text = extract_text(Path(d.path), 20000)
            if not text and d.extension not in (".pdf", ".docx", ".xlsx"):
                continue
            refs = set()
            for pat in REF_PATTERNS:
                for m in pat.finditer(text):
                    refs.add(m.group(m.lastindex))
            if d.extension == ".pdf":
                refs.update(self._pdf_links(Path(d.path)))
            if d.extension == ".docx":
                refs.update(self._docx_links(Path(d.path)))
            if d.extension == ".xlsx":
                refs.update(self._xlsx_external(Path(d.path)))

            for ref in refs:
                ref_path = ref.replace("file://", "").strip()
                ref_name = Path(ref_path).name.lower()
                target = path_map.get(Path(ref_path).resolve().as_posix().lower())
                if not target:
                    target = name_map.get(ref_name)
                if not target and ref_name:
                    for nm, desc in name_map.items():
                        if nm == ref_name or nm.endswith("/" + ref_name):
                            target = desc
                            break
                if target and target.file_id != d.file_id:
                    edges.append(
                        RelationEdge(
                            d.file_id, "REFERENCES", target.file_id, weight=0.6, props={"ref": ref}
                        )
                    )
        return edges

    def _pdf_links(self, path: Path) -> set[str]:
        links: set[str] = set()
        try:
            import fitz

            doc = fitz.open(path)
            for page in doc:
                for link in page.get_links():
                    uri = link.get("uri") or ""
                    if uri.startswith("file:"):
                        links.add(uri)
            doc.close()
        except Exception:
            pass
        return links

    def _docx_links(self, path: Path) -> set[str]:
        links: set[str] = set()
        try:
            from docx import Document
            from docx.opc.constants import RELATIONSHIP_TYPE as RT

            doc = Document(path)
            for rel in doc.part.rels.values():
                if rel.reltype == RT.HYPERLINK:
                    uri = rel.target_ref
                    if uri.startswith("file:"):
                        links.add(uri)
        except Exception:
            pass
        return links

    def _xlsx_external(self, path: Path) -> set[str]:
        links: set[str] = set()
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.hyperlink and cell.hyperlink.target:
                            t = cell.hyperlink.target
                            if "file:" in t or t.endswith((".xlsx", ".xls", ".csv")):
                                links.add(t)
            wb.close()
        except Exception:
            pass
        return links


class WeakFileParser(RelationParser):
    name = "weak_files"

    def discover(
        self, descriptors: list[FileDescriptor], store: Neo4jStore
    ) -> list[RelationEdge]:
        by_name = {d.name.lower(): d for d in descriptors}
        edges: list[RelationEdge] = []

        for d in descriptors:
            for pat, kind in TEMP_PATTERNS:
                m = pat.match(d.name)
                if not m:
                    continue
                host_name = m.group(1)
                host = by_name.get(host_name.lower()) or by_name.get(host_name)
                if host:
                    edges.append(
                        RelationEdge(
                            d.file_id, "IS_TEMPORARY_OF", host.file_id, weight=0.3
                        )
                    )

            for pat in BACKUP_PATTERNS:
                m = pat.match(d.name)
                if not m:
                    continue
                base_name = m.group(1)
                for ext in ("", ".xlsx", ".docx", ".pdf", ".csv"):
                    host = by_name.get(f"{base_name}{ext}".lower())
                    if host:
                        edges.append(
                            RelationEdge(
                                d.file_id, "IS_BACKUP_OF", host.file_id, weight=0.4
                            )
                        )
                        break
        return edges


class ContainsParser(RelationParser):
    name = "contains"

    def discover(
        self, descriptors: list[FileDescriptor], store: Neo4jStore
    ) -> list[RelationEdge]:
        edges: list[RelationEdge] = []
        name_map = {d.name.lower(): d for d in descriptors}
        for d in descriptors:
            if d.extension == ".zip":
                members = list_archive_members(Path(d.path))
                for member in members[:30]:
                    edges.append(
                        RelationEdge(
                            d.file_id,
                            "CONTAINS",
                            f"virtual:{d.file_id}:{member}",
                            weight=0.5,
                            props={"member": member, "virtual": True},
                        )
                    )
            if d.extension == ".pdf":
                for att in self._pdf_attachments(Path(d.path)):
                    target = name_map.get(Path(att).name.lower())
                    if target:
                        edges.append(
                            RelationEdge(
                                d.file_id,
                                "CONTAINS",
                                target.file_id,
                                weight=0.55,
                                props={"attachment": att},
                            )
                        )
                    else:
                        edges.append(
                            RelationEdge(
                                d.file_id,
                                "CONTAINS",
                                f"virtual:{d.file_id}:att:{att}",
                                weight=0.5,
                                props={"attachment": att, "virtual": True},
                            )
                        )
        return edges

    def _pdf_attachments(self, path: Path) -> list[str]:
        names: list[str] = []
        try:
            import fitz

            doc = fitz.open(path)
            for i in range(doc.embfile_count()):
                info = doc.embfile_info(i)
                names.append(info.get("filename") or f"embfile_{i}")
            doc.close()
        except Exception:
            pass
        return names
