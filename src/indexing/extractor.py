from __future__ import annotations

import re
import zipfile
from pathlib import Path

TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".py", ".js", ".ts", ".java", ".c", ".cpp",
    ".h", ".hpp", ".json", ".yaml", ".yml", ".xml", ".html", ".css", ".csv",
    ".log", ".ini", ".cfg", ".toml", ".rst", ".eml", ".ics",
}
DOC_EXTENSIONS = {".pdf", ".docx", ".xlsx"}


def extract_text(path: Path, max_chars: int = 8000) -> str:
    ext = path.suffix.lower()
    if ext in TEXT_EXTENSIONS:
        return _read_text(path, max_chars)
    if ext == ".pdf":
        return _extract_pdf(path, max_chars)
    if ext == ".docx":
        return _extract_docx(path, max_chars)
    if ext == ".xlsx":
        return _extract_xlsx(path, max_chars)
    return ""


def build_summary(path: Path, text: str) -> tuple[str, str]:
    """返回 (summary, ai_summary)。批量索引时跳过 LLM（FILEKG_INDEX_FAST）。"""
    import os

    from src.config import settings

    name = path.name
    parent = str(path.parent)
    ext = path.suffix

    if text:
        summary = text[:512].replace("\n", " ").strip()
        rule = _rule_summary(name, text)
        fast = os.environ.get("FILEKG_INDEX_FAST", "").lower() in ("1", "true", "yes")
        if fast or not settings.llm_enabled:
            ai = rule[:50]
        else:
            from src.llm.summarizer import generate_ai_summary

            ai = generate_ai_summary(name, text, rule_fallback=rule)
    else:
        summary = f"文件 {name}，位于 {parent}，类型 {ext or '未知'}"
        ai = summary[:50]

    return summary, ai[:50]


def _rule_summary(name: str, text: str) -> str:
    snippet = text[:2000].replace("\n", " ").strip()
    if len(snippet) < 20:
        return f"{name}：内容较短"
    first_sentence = re.split(r"[。！？\n]", snippet)[0][:80]
    return first_sentence or f"{name} 相关文档"


def _read_text(path: Path, max_chars: int) -> str:
    for enc in ("utf-8", "gbk", "latin-1"):
        try:
            return path.read_text(encoding=enc)[:max_chars]
        except (UnicodeDecodeError, OSError):
            continue
    return ""


def _extract_pdf(path: Path, max_chars: int) -> str:
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(path)
        parts = []
        for page in doc:
            parts.append(page.get_text())
            if sum(len(p) for p in parts) >= max_chars:
                break
        doc.close()
        return "\n".join(parts)[:max_chars]
    except Exception:
        return ""


def _extract_docx(path: Path, max_chars: int) -> str:
    try:
        from docx import Document

        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs)[:max_chars]
    except Exception:
        return ""


def _extract_xlsx(path: Path, max_chars: int) -> str:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(max_row=50, values_only=True):
                parts.append(" ".join(str(c) for c in row if c is not None))
        wb.close()
        return "\n".join(parts)[:max_chars]
    except Exception:
        return ""


def list_archive_members(archive_path: Path) -> list[str]:
    members: list[str] = []
    try:
        if archive_path.suffix.lower() == ".zip":
            with zipfile.ZipFile(archive_path) as zf:
                members = [n for n in zf.namelist() if not n.endswith("/")]
    except Exception:
        pass
    return members
